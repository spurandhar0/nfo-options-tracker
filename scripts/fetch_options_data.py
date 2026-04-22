"""
NSE F&O Stock Options Fetcher — FIXED for GitHub Actions
Data source : Yahoo Finance (yfinance) — no NSE API calls
"""

import yfinance as yf
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, date, timedelta
import calendar, os, json, time

OUT = os.path.join(os.path.dirname(__file__), "..", "data", "stock_options.xlsx")

NAVY="1F3864"; GOLD="C9A84C"; CE_BG="DDEEFF"; PE_BG="FFE8E8"
ATM_BG="FFFACD"; WHITE="FFFFFF"; LIGHT="F5F5F5"; GREEN="E2EFDA"; ORANGE="FCE4D6"
DATA_COLS = ["Prev Close","Open","High","Low","Close","OI","Chg OI","Volume","IV (%)"]

FO_STOCKS = [
    ("AARTIIND","AARTIIND.NS",1200,5),("ABB","ABB.NS",250,50),
    ("ABBOTINDIA","ABBOTINDIA.NS",40,200),("ABCAPITAL","ABCAPITAL.NS",3200,5),
    ("ABFRL","ABFRL.NS",2800,5),("ACC","ACC.NS",500,25),
    ("ADANIENT","ADANIENT.NS",875,50),("ADANIPORTS","ADANIPORTS.NS",1250,25),
    ("ALKEM","ALKEM.NS",200,100),("AMBUJACEM","AMBUJACEM.NS",2000,10),
    ("ANGELONE","ANGELONE.NS",300,50),("APOLLOHOSP","APOLLOHOSP.NS",125,100),
    ("APOLLOTYRE","APOLLOTYRE.NS",2700,10),("ASHOKLEY","ASHOKLEY.NS",5500,5),
    ("ASIANPAINT","ASIANPAINT.NS",200,50),("ASTRAL","ASTRAL.NS",400,25),
    ("ATGL","ATGL.NS",750,25),("AUROPHARMA","AUROPHARMA.NS",650,25),
    ("AUBANK","AUBANK.NS",1000,25),("AXISBANK","AXISBANK.NS",625,25),
    ("BAJAJ-AUTO","BAJAJ-AUTO.NS",250,100),("BAJAJFINSV","BAJAJFINSV.NS",500,25),
    ("BAJFINANCE","BAJFINANCE.NS",125,100),("BALKRISIND","BALKRISIND.NS",400,50),
    ("BANDHANBNK","BANDHANBNK.NS",3600,10),("BANKBARODA","BANKBARODA.NS",5850,5),
    ("BEL","BEL.NS",3750,5),("BERGEPAINT","BERGEPAINT.NS",1100,25),
    ("BHARTIARTL","BHARTIARTL.NS",500,25),("BHEL","BHEL.NS",10500,5),
    ("BIOCON","BIOCON.NS",2900,5),("BOSCHLTD","BOSCHLTD.NS",50,200),
    ("BPCL","BPCL.NS",1800,10),("BRITANNIA","BRITANNIA.NS",200,50),
    ("CANBK","CANBK.NS",4350,5),("CHOLAFIN","CHOLAFIN.NS",700,25),
    ("CIPLA","CIPLA.NS",650,25),("COALINDIA","COALINDIA.NS",2100,10),
    ("COFORGE","COFORGE.NS",200,100),("COLPAL","COLPAL.NS",350,25),
    ("CONCOR","CONCOR.NS",1250,25),("CROMPTON","CROMPTON.NS",3000,5),
    ("CUMMINSIND","CUMMINSIND.NS",600,50),("DABUR","DABUR.NS",2750,5),
    ("DEEPAKNTR","DEEPAKNTR.NS",300,50),("DIVISLAB","DIVISLAB.NS",200,100),
    ("DIXON","DIXON.NS",150,100),("DLF","DLF.NS",1650,10),
    ("DRREDDY","DRREDDY.NS",250,50),("EICHERMOT","EICHERMOT.NS",175,100),
    ("ESCORTS","ESCORTS.NS",275,50),("EXIDEIND","EXIDEIND.NS",3600,5),
    ("FEDERALBNK","FEDERALBNK.NS",5000,5),("GAIL","GAIL.NS",6400,5),
    ("GLENMARK","GLENMARK.NS",1150,25),("GODREJCP","GODREJCP.NS",1000,25),
    ("GODREJPROP","GODREJPROP.NS",400,50),("GRANULES","GRANULES.NS",2700,5),
    ("GRASIM","GRASIM.NS",475,25),("GUJGASLTD","GUJGASLTD.NS",750,25),
    ("HAL","HAL.NS",200,100),("HAVELLS","HAVELLS.NS",500,25),
    ("HCLTECH","HCLTECH.NS",350,25),("HDFCAMC","HDFCAMC.NS",200,100),
    ("HDFCBANK","HDFCBANK.NS",550,25),("HDFCLIFE","HDFCLIFE.NS",1500,10),
    ("HEROMOTOCO","HEROMOTOCO.NS",300,50),("HINDALCO","HINDALCO.NS",2150,10),
    ("HINDCOPPER","HINDCOPPER.NS",3650,5),("HINDPETRO","HINDPETRO.NS",2700,5),
    ("HINDUNILVR","HINDUNILVR.NS",300,50),("IDFCFIRSTB","IDFCFIRSTB.NS",7500,5),
    ("IEX","IEX.NS",3750,5),("IGL","IGL.NS",1375,10),
    ("INDHOTEL","INDHOTEL.NS",3000,5),("INDIAMART","INDIAMART.NS",150,100),
    ("INDIGO","INDIGO.NS",300,50),("INDUSINDBK","INDUSINDBK.NS",500,25),
    ("INDUSTOWER","INDUSTOWER.NS",2800,5),("INFY","INFY.NS",400,25),
    ("IOC","IOC.NS",3500,5),("IRCTC","IRCTC.NS",875,25),
    ("ITC","ITC.NS",3200,10),("JINDALSTEL","JINDALSTEL.NS",1250,25),
    ("JSWENERGY","JSWENERGY.NS",1500,10),("JSWSTEEL","JSWSTEEL.NS",600,25),
    ("JUBLFOOD","JUBLFOOD.NS",625,25),("KOTAKBANK","KOTAKBANK.NS",400,25),
    ("KPITTECH","KPITTECH.NS",800,25),("LALPATHLAB","LALPATHLAB.NS",300,100),
    ("LAURUSLABS","LAURUSLABS.NS",2500,5),("LICHSGFIN","LICHSGFIN.NS",1600,10),
    ("LT","LT.NS",150,50),("LTF","LTF.NS",5000,5),
    ("LTIM","LTIM.NS",150,100),("LTTS","LTTS.NS",200,100),
    ("LUPIN","LUPIN.NS",500,25),("M&M","M&M.NS",175,50),
    ("MANAPPURAM","MANAPPURAM.NS",4000,5),("MARICO","MARICO.NS",1200,10),
    ("MARUTI","MARUTI.NS",100,100),("MCX","MCX.NS",400,25),
    ("METROPOLIS","METROPOLIS.NS",400,50),("MFSL","MFSL.NS",1100,25),
    ("MGL","MGL.NS",550,25),("MOTHERSON","MOTHERSON.NS",14000,5),
    ("MPHASIS","MPHASIS.NS",300,50),("MRF","MRF.NS",5,500),
    ("MUTHOOTFIN","MUTHOOTFIN.NS",600,25),("NATIONALUM","NATIONALUM.NS",7500,5),
    ("NAUKRI","NAUKRI.NS",150,100),("NAVINFLUOR","NAVINFLUOR.NS",200,100),
    ("NESTLEIND","NESTLEIND.NS",50,100),("NMDC","NMDC.NS",6750,5),
    ("NTPC","NTPC.NS",2250,10),("OBEROIRLTY","OBEROIRLTY.NS",700,25),
    ("OFSS","OFSS.NS",75,100),("ONGC","ONGC.NS",1925,10),
    ("PAGEIND","PAGEIND.NS",15,500),("PEL","PEL.NS",750,25),
    ("PERSISTENT","PERSISTENT.NS",175,100),("PETRONET","PETRONET.NS",3000,10),
    ("PFC","PFC.NS",2700,5),("PIDILITIND","PIDILITIND.NS",250,50),
    ("PIIND","PIIND.NS",250,50),("PNB","PNB.NS",8000,5),
    ("POLYCAB","POLYCAB.NS",175,100),("POWERGRID","POWERGRID.NS",2700,10),
    ("RBLBANK","RBLBANK.NS",5000,5),("RECLTD","RECLTD.NS",3000,5),
    ("RELIANCE","RELIANCE.NS",1250,50),("SAIL","SAIL.NS",8500,5),
    ("SBICARD","SBICARD.NS",1000,25),("SBILIFE","SBILIFE.NS",750,25),
    ("SBIN","SBIN.NS",1500,10),("SHREECEM","SHREECEM.NS",25,500),
    ("SHRIRAMFIN","SHRIRAMFIN.NS",300,50),("SIEMENS","SIEMENS.NS",275,50),
    ("SRF","SRF.NS",375,50),("SUNPHARMA","SUNPHARMA.NS",350,25),
    ("SUNTV","SUNTV.NS",1500,10),("SYNGENE","SYNGENE.NS",1500,10),
    ("TATACHEM","TATACHEM.NS",1100,25),("TATACOMM","TATACOMM.NS",700,25),
    ("TATACONSUM","TATACONSUM.NS",1100,10),("TATAMOTORS","TATAMOTORS.NS",1400,10),
    ("TATAPOWER","TATAPOWER.NS",3375,5),("TATASTEEL","TATASTEEL.NS",5500,5),
    ("TCS","TCS.NS",150,50),("TECHM","TECHM.NS",400,25),
    ("TITAN","TITAN.NS",375,50),("TORNTPHARM","TORNTPHARM.NS",500,50),
    ("TORNTPOWER","TORNTPOWER.NS",1500,25),("TRENT","TRENT.NS",375,50),
    ("TVSMOTOR","TVSMOTOR.NS",350,25),("UBL","UBL.NS",700,25),
    ("ULTRACEMCO","ULTRACEMCO.NS",200,100),("UNIONBANK","UNIONBANK.NS",8400,5),
    ("UPL","UPL.NS",1300,25),("VEDL","VEDL.NS",2750,10),
    ("VOLTAS","VOLTAS.NS",1000,25),("WIPRO","WIPRO.NS",1500,10),
    ("ZOMATO","ZOMATO.NS",4500,5),("ZYDUSLIFE","ZYDUSLIFE.NS",700,25),
]

def thin():
    s=Side(style="thin",color="CCCCCC"); return Border(left=s,right=s,top=s,bottom=s)

def hdr(cell,bg=NAVY,fg="FFFFFF",sz=9,bold=True):
    cell.font=Font(name="Arial",bold=bold,color=fg,size=sz)
    cell.fill=PatternFill("solid",start_color=bg)
    cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
    cell.border=thin()

def vc(ws,row,col,value,bg=WHITE,fmt=None,bold=False):
    c=ws.cell(row=row,column=col,value=value)
    c.font=Font(name="Arial",size=9,bold=bold)
    c.fill=PatternFill("solid",start_color=bg)
    c.alignment=Alignment(horizontal="center",vertical="center")
    c.border=thin()
    if fmt: c.number_format=fmt
    return c

def last_thursday(year,month):
    last_day=calendar.monthrange(year,month)[1]
    d=date(year,month,last_day)
    return d-timedelta(days=(d.weekday()-3)%7)

def get_expiries():
    today=date.today()
    cm=last_thursday(today.year,today.month)
    if today>cm:
        nm=today.month%12+1; ny=today.year+(1 if today.month==12 else 0)
        cm=last_thursday(ny,nm)
        nm2=nm%12+1; ny2=ny+(1 if nm==12 else 0)
        nm_exp=last_thursday(ny2,nm2)
    else:
        nm=today.month%12+1; ny=today.year+(1 if today.month==12 else 0)
        nm_exp=last_thursday(ny,nm)
    return cm,nm_exp

def round_to_interval(price,interval):
    return int(round(price/interval)*interval)

def closest_expiry(available,target):
    best=None; best_diff=999
    for s in available:
        try:
            d=datetime.strptime(s,"%Y-%m-%d").date()
            diff=abs((d-target).days)
            if diff<best_diff: best_diff=diff; best=s
        except: pass
    return best

def fetch_chain(yf_sym,target):
    try:
        tk=yf.Ticker(yf_sym)
        avail=tk.options
        if not avail: return {},None
        exp=closest_expiry(list(avail),target)
        if not exp: return {},None
        ch=tk.option_chain(exp)
        result={}
        def parse(df,otype):
            for _,row in df.iterrows():
                strike=float(row.get("strike",0))
                if strike not in result: result[strike]={"CE":{},"PE":{}}
                result[strike][otype]={
                    "Prev Close":round(float(row.get("lastPrice",0) or 0),2),
                    "Open":0,
                    "High":round(float(row.get("highPrice",row.get("lastPrice",0)) or 0),2),
                    "Low":round(float(row.get("lowPrice",row.get("lastPrice",0)) or 0),2),
                    "Close":round(float(row.get("lastPrice",0) or 0),2),
                    "OI":int(row.get("openInterest",0) or 0),
                    "Chg OI":0,
                    "Volume":int(row.get("volume",0) or 0),
                    "IV (%)":round(float(row.get("impliedVolatility",0) or 0)*100,2),
                }
        parse(ch.calls,"CE"); parse(ch.puts,"PE")
        return result,exp
    except Exception as e:
        return {},None

def write_sheet(wb,nse_sym,spot,opt_data,atm,interval,exp_label,suffix):
    sname=f"{nse_sym}_{suffix}"
    if sname in wb.sheetnames: del wb[sname]
    ws=wb.create_sheet(sname)
    ws.sheet_view.showGridLines=False
    wk_label="Current Month" if suffix=="CM" else "Next Month"
    total_cols=len(DATA_COLS)+1+len(DATA_COLS)
    ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=total_cols)
    t=ws.cell(row=1,column=1,
        value=f"{nse_sym}  |  Spot: ₹{spot:,.2f}  |  Expiry: {exp_label}  ({wk_label})  |  {datetime.now().strftime('%d-%b-%Y')}")
    t.font=Font(name="Arial",bold=True,size=11,color="FFFFFF")
    t.fill=PatternFill("solid",start_color=NAVY)
    t.alignment=Alignment(horizontal="center",vertical="center")
    ws.row_dimensions[1].height=26
    ws.merge_cells(start_row=2,start_column=1,end_row=2,end_column=len(DATA_COLS))
    hdr(ws.cell(row=2,column=1,value="CALL (CE)"),bg="1A5276",sz=10)
    sc=len(DATA_COLS)+1
    hdr(ws.cell(row=2,column=sc,value="STRIKE"),bg=NAVY,sz=10)
    ps=sc+1
    ws.merge_cells(start_row=2,start_column=ps,end_row=2,end_column=ps+len(DATA_COLS)-1)
    hdr(ws.cell(row=2,column=ps,value="PUT (PE)"),bg="922B21",sz=10)
    ws.row_dimensions[2].height=22
    for ci,col in enumerate(DATA_COLS,1):
        ws.cell(row=3,column=ci,value=col); hdr(ws.cell(row=3,column=ci),bg="1F618D",sz=8)
    ws.cell(row=3,column=sc,value="Strike"); hdr(ws.cell(row=3,column=sc),bg=GOLD,fg=NAVY,sz=9)
    for ci,col in enumerate(DATA_COLS,ps):
        ws.cell(row=3,column=ci,value=col); hdr(ws.cell(row=3,column=ci),bg="922B21",sz=8)
    ws.row_dimensions[3].height=32; ws.freeze_panes="A4"
    strikes=sorted([atm+i*interval for i in range(-6,7)],reverse=True)
    fmts=["#,##0.00"]*5+["#,##0","#,##0","#,##0","#,##0.00"]
    for ri,strike in enumerate(strikes,4):
        is_atm=(strike==atm)
        ce_bg=ATM_BG if is_atm else CE_BG; pe_bg=ATM_BG if is_atm else PE_BG
        row_bg=ATM_BG if is_atm else WHITE
        ce=opt_data.get(float(strike),{}).get("CE",{}); pe=opt_data.get(float(strike),{}).get("PE",{})
        for ci,(col,fmt) in enumerate(zip(DATA_COLS,fmts),1): vc(ws,ri,ci,ce.get(col,0) or 0,ce_bg,fmt)
        c=vc(ws,ri,sc,strike,row_bg,"#,##0",bold=True)
        if is_atm: c.font=Font(name="Arial",bold=True,size=10,color="8B0000")
        for ci,(col,fmt) in enumerate(zip(DATA_COLS,fmts),ps): vc(ws,ri,ci,pe.get(col,0) or 0,pe_bg,fmt)
    for ci in range(1,total_cols+1):
        ws.column_dimensions[get_column_letter(ci)].width=9 if ci==sc else 11

def write_summary(wb,stocks_done,cm_exp,nm_exp):
    sname="📊 Summary"
    if sname in wb.sheetnames: del wb[sname]
    ws=wb.create_sheet(sname,0); ws.sheet_view.showGridLines=False
    today=datetime.now().strftime("%d-%b-%Y %H:%M")
    ws.merge_cells("A1:G1")
    t=ws.cell(row=1,column=1,
        value=f"NSE Stock Options  |  CM:{cm_exp.strftime('%d-%b-%Y')}  NM:{nm_exp.strftime('%d-%b-%Y')}  |  {today}")
    t.font=Font(name="Arial",bold=True,size=12,color="FFFFFF")
    t.fill=PatternFill("solid",start_color=NAVY)
    t.alignment=Alignment(horizontal="center",vertical="center"); ws.row_dimensions[1].height=30
    for ci,h in enumerate(["#","Symbol","Spot (₹)","Lot Size","Interval","ATM","Sheets"],1):
        ws.cell(row=2,column=ci,value=h); hdr(ws.cell(row=2,column=ci),bg=GOLD,fg=NAVY)
    ws.row_dimensions[2].height=20
    for ri,s in enumerate(stocks_done,3):
        bg=WHITE if ri%2==0 else LIGHT
        vc(ws,ri,1,ri-2,bg); vc(ws,ri,2,s["sym"],bg,bold=True)
        vc(ws,ri,3,s["spot"],bg,"#,##0.00"); vc(ws,ri,4,s["lot"],bg)
        vc(ws,ri,5,s["interval"],bg); vc(ws,ri,6,s["atm"],bg,"#,##0")
        vc(ws,ri,7,f"{s['sym']}_CM | {s['sym']}_NM",bg)
    for col,w in zip(["A","B","C","D","E","F","G"],[5,14,14,10,10,14,22]):
        ws.column_dimensions[col].width=w
    ws.freeze_panes="A3"; ws.auto_filter.ref="A2:G2"

def write_legend(wb):
    sname="📋 Legend"
    if sname in wb.sheetnames: del wb[sname]
    ws=wb.create_sheet(sname); ws.sheet_view.showGridLines=False
    ws.merge_cells("A1:B1")
    t=ws.cell(row=1,column=1,value="Legend & Column Guide")
    t.font=Font(name="Arial",bold=True,size=13,color="FFFFFF")
    t.fill=PatternFill("solid",start_color=NAVY); t.alignment=Alignment(horizontal="center",vertical="center")
    ws.row_dimensions[1].height=28
    legends=[
        ("Sheet Naming","SYMBOL_CM = Current Month | SYMBOL_NM = Next Month"),
        ("Monthly Expiry","Last Thursday of each month (NSE standard for stock F&O)"),
        ("ATM Row","Yellow = At-The-Money strike (nearest to spot price)"),
        ("CALL (CE)","Blue — buyer profits when stock RISES above strike"),
        ("PUT (PE)","Red — buyer profits when stock FALLS below strike"),
        ("Close","Last traded price of the option contract"),
        ("OI","Open Interest — total outstanding contracts"),
        ("Volume","Contracts traded in current session"),
        ("IV (%)","Implied Volatility (%) from Yahoo Finance"),
        ("Open / Chg OI","0 = not available via Yahoo Finance option chain"),
        ("Data Source","Yahoo Finance (yfinance) — NSE option chain"),
        ("Schedule","GitHub Actions — Every weekday Mon–Fri at 6:30 PM IST"),
    ]
    for ri,(term,desc) in enumerate(legends,3):
        c1=ws.cell(row=ri,column=1,value=term); c1.font=Font(name="Arial",bold=True,size=9)
        c1.fill=PatternFill("solid",start_color=LIGHT); c1.border=thin()
        c2=ws.cell(row=ri,column=2,value=desc); c2.font=Font(name="Arial",size=9); c2.border=thin()
        ws.row_dimensions[ri].height=16
    ws.column_dimensions["A"].width=22; ws.column_dimensions["B"].width=70

def main():
    print("🚀 NSE F&O Stock Options Tracker (via Yahoo Finance)")
    cm_exp,nm_exp=get_expiries()
    print(f"   CM Expiry: {cm_exp}  |  NM Expiry: {nm_exp}")
    print(f"   Stocks: {len(FO_STOCKS)}")
    os.makedirs(os.path.dirname(os.path.abspath(OUT)),exist_ok=True)
    wb=Workbook(); wb.remove(wb.active)
    stocks_done=[]; failed=[]
    for i,(nse_sym,yf_sym,lot,interval) in enumerate(FO_STOCKS,1):
        print(f"  [{i:>3}/{len(FO_STOCKS)}] {nse_sym:<15}",end="  ")
        try:
            tk=yf.Ticker(yf_sym)
            hist=tk.history(period="3d")
            if hist.empty: print("⚠️ skip"); failed.append(nse_sym); continue
            spot=round(float(hist["Close"].iloc[-1]),2)
            atm=round_to_interval(spot,interval)
            print(f"Spot={spot:>10,.2f}  ATM={atm:>8,}",end="  ")
            cm_data,cm_lbl=fetch_chain(yf_sym,cm_exp)
            nm_data,nm_lbl=fetch_chain(yf_sym,nm_exp)
            if cm_data: write_sheet(wb,nse_sym,spot,cm_data,atm,interval,cm_lbl or str(cm_exp),"CM")
            if nm_data: write_sheet(wb,nse_sym,spot,nm_data,atm,interval,nm_lbl or str(nm_exp),"NM")
            stocks_done.append({"sym":nse_sym,"spot":spot,"lot":lot,"interval":interval,"atm":atm})
            print("✅")
        except Exception as e:
            print(f"❌ {e}"); failed.append(nse_sym)
        time.sleep(0.3)
    write_summary(wb,stocks_done,cm_exp,nm_exp)
    write_legend(wb)
    out=os.path.abspath(OUT); wb.save(out)
    print(f"\n✅ Saved: {out}  |  OK:{len(stocks_done)}  Fail:{len(failed)}")
    if failed: print(f"   Failed: {', '.join(failed)}")
    with open(out.replace(".xlsx",".json"),"w") as f:
        json.dump({"updated_at":datetime.now().isoformat(),"cm":str(cm_exp),
                   "nm":str(nm_exp),"total":len(stocks_done),"failed":failed},f,indent=2)
    print("🎉 Done!")

if __name__=="__main__":
    main()
